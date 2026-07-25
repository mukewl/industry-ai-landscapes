"""Generate skeleton landscape workbooks for gaming / fmcg / luxury.

One IDENTICAL structural skeleton for all three industries (same column
positions, same prefix conventions w1_/d1_/ai1_/vc1_/f1_) so a single
extractor and data contract serve every industry; the industry-specific
MEANING lives in each workbook's Scoring Guide sheet (generated from the
INDUSTRIES dict below) and in industries/<ind>/frame.md.

Derived scores (W%/D%/AI%/EPI/tier/quadrant) are written as PLAIN VALUES by
scripts/populate.py at population time — deliberately NOT Excel formulas, so
openpyxl extraction never hits the cached-formula-value trap that bit the
travel project.

    python -X utf8 scripts/make_workbook.py            # writes all three
    python -X utf8 scripts/make_workbook.py gaming     # just one
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
SCHEMA_VERSION = "1.0"

# ---------------------------------------------------------------- shared skeleton
# (phase, header) — headers snake_case to the keys documented in template/DATA-CONTRACT.md.
# Prefix conventions the machinery matches on: w1_..w5_, d1_..d6_, ai1_..ai7_,
# vc1_..vc7_ (value-chain stages), f1_..f5_ (financial health).
PROFILE = [
    "Company", "HQ", "Founded", "FTEs", "Total funding mn", "Market cap valuation mn",
    "Source sector taxonomy", "Current sector focus", "Business model notes",
    "Revenue by product", "Customer base segment", "Merchant of record",
    "Building infrastructure in house", "Legacy rail dependency", "API commerce ready",
    "Direct channel", "MCP", "A2A", "UCP", "ACP",
    "AI capabilities", "Is AI the core focus", "Foundation model approach",
    "Proprietary LLM on industry data", "Customer facing AI assistant",
    "Agentic AI browse buy pay", "Agentic maturity layer", "AI personalisation engine",
    "Proprietary data advantage", "Gen AI platform partnerships", "AI governance safety",
    "Existing partnerships", "Industry fingerprint",
]
POSTURE = [
    "Residual gap what they d need", "Suggested action", "Threat tier", "Horizon",
    "Impact on incumbent line", "Final action", "Team notes", "Links",
    "Reviewer notes", "Last reviewed", "Owner",
]
ELIMINATION = [
    "Last round and date", "Post money valuation mn", "Revenue traction",
    "Burn profitability", "Cash and runway est", "Debt leverage",
    "F1 funding adequacy and recency", "F2 revenue and profitability", "F3 runway",
    "F4 leverage 5 low debt", "F5 capital access backers",
    "Financial health pct", "Survival tier", "Survival outlook", "Solvency ratio pct",
]
EVIDENCE = [
    "Evidence links", "Evidence notes", "Confidence", "Researched on", "Researched by",
]
POTENTIAL_COMPUTED = [
    "Willingness pct", "Distribution readiness pct", "AI readiness pct",
    "Readiness pct combined", "Entry potential index", "Entry tier", "Quadrant",
]

# ------------------------------------------------------- per-industry definitions
# Every signal: (short header suffix, definition, anchor examples "0 · 3 · 5").
# AI1–AI7 are structurally identical across industries (the AI stack IS the
# cross-industry lens); W/D and the vocabulary differ per frame.md.
AI_SIGNALS = [
    ("Foundation models", "Depth of foundation-model capability (none → API user → fine-tunes → trains own)",
     "0 no AI · 3 builds seriously on external models · 5 trains frontier/proprietary models"),
    ("Agentic capability L1 3", "Autonomy of shipped agents: L1 assistive → L2 semi-autonomous → L3 end-to-end act-and-pay",
     "0 none · 3 L2 in production · 5 L3 browse/transact/pay live"),
    ("Agentic commerce protocols", "Adoption of MCP / A2A / UCP / ACP or equivalent agent-to-commerce standards",
     "0 none · 3 one protocol piloted · 5 multiple protocols in production"),
    ("Conversational assistant", "Customer-facing assistant quality and adoption",
     "0 none · 3 useful assistant shipped · 5 assistant is a primary interface"),
    ("Personalisation engine", "AI personalisation of offers/content per user",
     "0 none · 3 segment-level · 5 individual-level, core to the product"),
    ("Proprietary data advantage", "Unique behavioral/transaction data others cannot get",
     "0 commodity data · 3 useful proprietary set · 5 industry-defining data moat"),
    ("Gen AI ecosystem position", "Position in the gen-AI value chain (dependent → partner → platform others depend on)",
     "0 pure consumer of AI · 3 strategic AI partnerships · 5 others build on its AI"),
]

INDUSTRIES = {
    "gaming": {
        "title": "Gaming — AI/agentic disruption of game discovery & distribution",
        "incumbent": "The storefront take-rate model: Steam (Valve), console stores (PlayStation/Xbox/Nintendo), mobile app stores (Apple/Google)",
        "verticals": ["PC", "Console", "Mobile", "Cloud streaming", "UGC platforms"],
        "stages": ["Discover", "Evaluate", "Acquire", "Pay", "Play", "Engage", "Retain"],
        "w": [
            ("Strategic intent stance", "Public strategy/statements toward disrupting game distribution or storefront economics",
             "0 no interest · 3 exploring own channel · 5 openly attacking store take-rates"),
            ("Investment and M&A direction", "Capital deployed toward distribution/discovery assets (stores, launchers, community platforms)",
             "0 none · 3 one relevant bet · 5 systematic portfolio (e.g. Epic-style)"),
            ("Gaming partnerships and alliances", "Alliances with studios/publishers/platforms that reposition distribution",
             "0 none · 3 a few content deals · 5 load-bearing alliances across the stack"),
            ("Adjacent offerings shipped", "Launcher, store, subscription, cloud or payment experiments already live",
             "0 none · 3 one shipped experiment · 5 full storefront/subscription live"),
            ("Owned audience with gaming intent", "Direct reach to players who would follow it to a new channel",
             "0 no audience · 3 large adjacent audience · 5 massive player base with purchase intent"),
        ],
        "d": [
            ("Catalog and content access", "Games/IP it owns or can distribute (first-party or aggregated)",
             "0 none · 3 meaningful catalog or aggregation deals · 5 must-have exclusive catalog"),
            ("Distribution infrastructure", "Launcher/storefront tech, CDN, patching, anti-cheat, cloud delivery",
             "0 none · 3 partial stack · 5 proven store-grade infrastructure at scale"),
            ("Platform policy and compliance", "Ability to navigate ratings, regional regulation, and platform-holder rules",
             "0 naive · 3 competent in one region/segment · 5 global compliance machine"),
            ("Payments and merchant of record", "Takes the payment, handles tax/refunds/chargebacks itself",
             "0 no payments · 3 payments via partner · 5 global MoR at scale"),
            ("Global scale and reach", "Geographic + demographic reach of its channel today",
             "0 niche · 3 multi-region · 5 global, hundreds of millions"),
            ("Brand trust with players", "Would players buy games from it tomorrow?",
             "0 unknown/distrusted · 3 respected in a niche · 5 beloved global gaming brand"),
        ],
        "protocol_note": "Direct channel = own launcher/store/web-shop bypassing platform storefronts (the industry's NDC-equivalent).",
    },
    "fmcg": {
        "title": "FMCG — agentic commerce bypassing the retail shelf",
        "incumbent": "The brand × big-retail model: CPG giants (P&G, Unilever, Nestlé) selling through the physical + e-commerce shelf (Walmart, Tesco, Amazon) and retail media",
        "verticals": ["Food and beverage", "Home care", "Personal care and beauty", "Health and wellness", "Pet"],
        "stages": ["Awareness", "Consideration", "Purchase", "Fulfilment", "Consumption", "Replenishment", "Loyalty"],
        "w": [
            ("Strategic intent stance", "Public strategy toward bypassing the retail shelf / owning the consumer directly",
             "0 shelf-loyal · 3 DTC experiments · 5 explicit direct-to-consumer or agent-channel strategy"),
            ("Investment and M&A direction", "Capital toward commerce, subscription, fulfilment or agentic-checkout assets",
             "0 none · 3 one relevant bet · 5 systematic portfolio"),
            ("Commerce partnerships and alliances", "Alliances (platforms, wallets, delivery, AI assistants) that reroute the purchase",
             "0 none · 3 a few pilots · 5 load-bearing alliances"),
            ("Adjacent offerings shipped", "DTC stores, subscriptions, replenishment programs, shoppable media already live",
             "0 none · 3 one shipped · 5 full direct commerce loop live"),
            ("Owned consumer relationship", "Direct, repeat consumer reach (app/subscription/community) with purchase intent",
             "0 none (shelf-anonymous) · 3 meaningful DTC base · 5 daily consumer relationship at scale"),
        ],
        "d": [
            ("Product and assortment access", "Owns or aggregates the products/brands consumers actually want",
             "0 none · 3 strong niche assortment · 5 full-basket assortment or must-have brands"),
            ("Fulfilment and logistics", "Warehousing, last-mile, cold-chain, returns — the physical rail",
             "0 none · 3 partner-dependent · 5 owned national/global fulfilment at scale"),
            ("Regulatory and compliance", "Food safety, labeling, claims, cross-border — competence to sell consumables",
             "0 naive · 3 competent in one market · 5 global compliance machine"),
            ("Payments and merchant of record", "Owns checkout, subscription billing, tax and refunds",
             "0 none · 3 via partner · 5 global MoR with one-click/replenishment billing"),
            ("Global scale and reach", "Market coverage of its channel today",
             "0 niche · 3 multi-region · 5 global mass reach"),
            ("Brand trust for the weekly basket", "Would consumers trust it with routine household purchases?",
             "0 unknown · 3 trusted in a category · 5 default household name"),
        ],
        "protocol_note": "Direct channel = DTC/subscription/agent-checkout route that skips the retail shelf (the industry's NDC-equivalent).",
    },
    "luxury": {
        "title": "Luxury — AI vs the controlled-distribution maison model",
        "incumbent": "The maison model: LVMH/Kering/Richemont-style houses controlling scarcity, price and the client relationship through boutiques and tightly-policed wholesale",
        "verticals": ["Fashion and leather", "Watches and jewelry", "Beauty and fragrance", "Wines and spirits", "Hospitality and experiences"],
        "stages": ["Inspire", "Discover", "Consider", "Purchase", "Own", "Aftercare", "Resale"],
        "w": [
            ("Strategic intent stance", "Public strategy toward re-intermediating luxury discovery, purchase or resale",
             "0 none · 3 experimenting at the edges · 5 explicit challenge to maison-controlled distribution"),
            ("Investment and M&A direction", "Capital toward luxury commerce, resale, authentication or clienteling assets",
             "0 none · 3 one relevant bet · 5 systematic portfolio"),
            ("Luxury partnerships and alliances", "Brand/retail/platform alliances that reposition who owns the client",
             "0 none · 3 a few pilots · 5 load-bearing alliances with major houses"),
            ("Adjacent offerings shipped", "Shoppable content, personal-shopper AI, resale/authentication, VIP programs live",
             "0 none · 3 one shipped · 5 full client-journey offering live"),
            ("Owned affluent clientele", "Direct reach to high-net-worth / aspirational clients with luxury intent",
             "0 none · 3 meaningful affluent audience · 5 the client list every maison wants"),
        ],
        "d": [
            ("Product and brand access", "Access to genuine luxury supply (brand partnerships, consignment, grey-market-free)",
             "0 none · 3 partial/selective access · 5 broad authorized access or own coveted brands"),
            ("Experience and service infrastructure", "Clienteling, white-glove fulfilment, boutiques/concessions, authentication",
             "0 none · 3 digital-only competence · 5 full omnichannel luxury-grade service"),
            ("Compliance and brand-equity discipline", "Ability to meet houses' price/presentation/selective-distribution rules",
             "0 naive/grey · 3 accepted by some brands · 5 trusted steward of brand equity"),
            ("Payments and merchant of record", "High-value checkout, financing, cross-border duties, fraud",
             "0 none · 3 via partner · 5 global high-ticket MoR"),
            ("Global scale and reach", "Reach across luxury capitals and growth markets",
             "0 single market · 3 multi-region · 5 global incl. APAC/Gulf luxury hubs"),
            ("Prestige and trust", "Would a maison let it touch their brand; would clients buy a Birkin through it?",
             "0 no prestige · 3 credible premium player · 5 unquestioned luxury credibility"),
        ],
        "protocol_note": "Direct channel = a client-owned route (app/AI stylist/resale platform) outside maison-controlled distribution (the industry's NDC-equivalent).",
    },
}

MODEL_CONFIG = [
    ("MODEL CONFIG — weights & thresholds (identical to the reference travel model; change only with a DECISIONS.md entry)", None),
    ("Entry-Potential blend", None),
    ("Willingness weight", 0.4),
    ("Readiness weight", 0.6),
    ("Readiness blend", None),
    ("Distribution readiness weight", 0.5),
    ("AI readiness weight", 0.5),
    ("Signal weights — Willingness W1..W5", "0.25 / 0.25 / 0.20 / 0.20 / 0.10"),
    ("Signal weights — Distribution D1..D6", "0.25 / 0.25 / 0.15 / 0.15 / 0.10 / 0.10"),
    ("Signal weights — AI AI1..AI7", "0.20 / 0.25 / 0.15 / 0.10 / 0.10 / 0.15 / 0.05"),
    ("Band thresholds", None),
    ("High band cut (>= High)", 60),
    ("Medium band cut (>= Medium)", 40),
    ("Quadrants", "Imminent threat (W>=60, R>=60) · Aspirant (W>=60, R<60) · Sleeping giant (W<60, R>=60) · Dormant"),
    ("Derived scores are computed by scripts/populate.py as plain values (no Excel formulas).", None),
    ("Schema version", SCHEMA_VERSION),
]

HDR_FILL = PatternFill("solid", fgColor="1F2A44")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
PHASE_FILL = PatternFill("solid", fgColor="0E1626")
PHASE_FONT = Font(color="7FD7FF", bold=True, size=11)


def columns_for(ind):
    """Full ordered header list + phase spans for one industry."""
    c = INDUSTRIES[ind]
    cols, phases = [], []

    def block(phase, headers):
        phases.append((phase, len(cols) + 1, len(cols) + len(headers)))
        cols.extend(headers)

    profile = PROFILE + [f"VC{i+1} {s}" for i, s in enumerate(c["stages"])]
    block("PHASE 1 · PROFILE", profile)
    position = []
    for v in c["verticals"]:
        position += [f"{v} distributing", f"{v} role"]
    position += ["Num verticals distributed", "Position class", "Motion",
                 "Target vertical s", "Phase 2 rationale"]
    block("PHASE 2 · POSITION", position)
    potential = [f"W{i+1} {name}" for i, (name, _, _) in enumerate(c["w"])]
    potential += [f"D{i+1} {name}" for i, (name, _, _) in enumerate(c["d"])]
    potential += [f"AI{i+1} {name}" for i, (name, _, _) in enumerate(AI_SIGNALS)]
    potential += POTENTIAL_COMPUTED
    block("PHASE 3 · POTENTIAL", potential)
    block("PHASE 4 · POSTURE", POSTURE)
    block("PHASE 5 · ELIMINATION", ELIMINATION)
    block("EVIDENCE", EVIDENCE)
    return cols, phases


def build(ind):
    c = INDUSTRIES[ind]
    cols, phases = columns_for(ind)
    wb = Workbook()

    ws = wb.active
    ws.title = "Landscape"
    for phase, a, b in phases:                       # row 1: phase banners
        ws.cell(row=1, column=a, value=phase).font = PHASE_FONT
        for col in range(a, b + 1):
            ws.cell(row=1, column=col).fill = PHASE_FILL
    ws.cell(row=2, column=1, value=f"{c['title']}  ·  headers in row 3, data from row 4  ·  schema v{SCHEMA_VERSION}")
    for j, h in enumerate(cols, 1):                  # row 3: headers
        cell = ws.cell(row=3, column=j, value=h)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(h) + 2))
    ws.freeze_panes = "B4"

    mc = wb.create_sheet("Model Config")
    for i, (label, val) in enumerate(MODEL_CONFIG, 1):
        mc.cell(row=i, column=1, value=label)
        if val is not None:
            mc.cell(row=i, column=2, value=val)
    mc.column_dimensions["A"].width = 60

    sg = wb.create_sheet("Scoring Guide")
    sg.append([f"SCORING GUIDE — {c['title']}"])
    sg.append([f"Incumbent under threat: {c['incumbent']}"])
    sg.append([c["protocol_note"]])
    sg.append([])
    sg.append(["Signal", "Definition", "Anchors (0 · 3 · 5)"])
    for i, (name, definition, anchors) in enumerate(c["w"]):
        sg.append([f"W{i+1} {name}", definition, anchors])
    for i, (name, definition, anchors) in enumerate(c["d"]):
        sg.append([f"D{i+1} {name}", definition, anchors])
    for i, (name, definition, anchors) in enumerate(AI_SIGNALS):
        sg.append([f"AI{i+1} {name}", definition, anchors])
    sg.append([])
    sg.append(["Value-chain stages (VC1..VC7)", " → ".join(c["stages"])])
    sg.append(["Verticals (Phase 2)", " · ".join(c["verticals"])])
    for col, width in (("A", 42), ("B", 70), ("C", 70)):
        sg.column_dimensions[col].width = width

    out = ROOT / "industries" / ind / f"{ind}_landscape.xlsx"
    wb.save(out)
    print(f"{out.relative_to(ROOT)}: {len(cols)} columns, sheets {wb.sheetnames}")
    return out


if __name__ == "__main__":
    targets = sys.argv[1:] or list(INDUSTRIES)
    for ind in targets:
        if ind not in INDUSTRIES:
            sys.exit(f"unknown industry {ind!r} — choose from {list(INDUSTRIES)}")
        build(ind)
