"""Quality gate for a populated industry workbook.

    python -X utf8 scripts/qa_check.py igaming

Run after every populate. Prints PASS/WARN/FAIL per check and exits non-zero on any
FAIL so a batch never silently degrades the dataset. Checks: schema integrity,
evidence coverage, score ranges, EPI recomputation, duplicates, vocabulary,
distribution sanity, and calibration-anchor drift.
"""
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
from populate import AIW, DW, FW, HEADER_ROW, DATA_START_ROW, WW, band, pct, quadrant, snake  # noqa: E402

ANCHOR_TOLERANCE = 8  # EPI points a recorded anchor may drift before it's a FAIL

results = []


def check(name, ok, detail="", warn=False):
    status = "PASS" if ok else ("WARN" if warn else "FAIL")
    results.append((status, name, detail))
    print(f"[{status}] {name}" + (f" — {detail}" if detail else ""))


def load(industry):
    wb = openpyxl.load_workbook(ROOT / "industries" / industry / f"{industry}_landscape.xlsx", data_only=True)
    ws = wb["Landscape"]
    headers = [snake(c.value) if c.value else None for c in ws[HEADER_ROW]]
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        rows.append({h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers) if h})
    return wb, rows


def sig(row, prefix, n):
    out = []
    for i in range(1, n + 1):
        key = next((k for k in row if k.startswith(f"{prefix}{i}_")), None)
        out.append(row.get(key) if key else None)
    return out


def main():
    industry = sys.argv[1] if len(sys.argv) > 1 else "igaming"
    wb, rows = load(industry)
    print(f"\n=== QA · {industry} · {len(rows)} companies ===\n")
    if not rows:
        check("populated rows", False, "workbook is empty")
        sys.exit(1)

    # 1. evidence coverage
    with_ev = [r for r in rows if r.get("evidence_links")]
    check("evidence coverage", len(with_ev) == len(rows),
          f"{len(with_ev)}/{len(rows)} rows cite sources")

    # 2. score ranges + completeness
    bad_range, missing_sig = [], []
    for r in rows:
        w, d, ai = sig(r, "w", 5), sig(r, "d", 6), sig(r, "ai", 7)
        for label, vec in (("W", w), ("D", d), ("AI", ai)):
            for i, v in enumerate(vec, 1):
                if v is None:
                    continue
                if not isinstance(v, (int, float)) or not (0 <= v <= 5):
                    bad_range.append(f"{r['company']} {label}{i}={v!r}")
        if sum(1 for v in w + d + ai if v is None) > 4:
            missing_sig.append(str(r.get("company")))
    check("signal scores in 0-5", not bad_range, "; ".join(bad_range[:5]))
    check("signal completeness", not missing_sig,
          f"{len(missing_sig)} rows with >4 null signals: {', '.join(missing_sig[:5])}", warn=True)

    # 3. EPI recomputation
    drift = []
    for r in rows:
        w = pct(sig(r, "w", 5), WW)
        d = pct(sig(r, "d", 6), DW)
        ai = pct(sig(r, "ai", 7), AIW)
        if None in (w, d, ai):
            continue
        rr = round(.5 * d + .5 * ai)
        epi = round(.4 * w + .6 * rr)
        stored = r.get("entry_potential_index")
        if stored is not None and abs(float(stored) - epi) > 1:
            drift.append(f"{r['company']}: stored {stored} vs recomputed {epi}")
        q = quadrant(w, rr)
        if r.get("quadrant") and q and r["quadrant"] != q:
            drift.append(f"{r['company']}: quadrant {r['quadrant']} vs {q}")
    check("EPI/quadrant recomputation", not drift, "; ".join(drift[:5]))

    # 4. duplicates
    names = [str(r["company"]).strip().lower() for r in rows]
    dupes = [n for n, c in Counter(names).items() if c > 1]
    near = []
    simple = [re.sub(r"\b(group|plc|ab|inc|ltd|limited|holdings|entertainment|nv|sa)\b|[^a-z0-9]", "", n) for n in names]
    for n, c in Counter(simple).items():
        if n and c > 1:
            near.append(n)
    check("no duplicate companies", not dupes, ", ".join(dupes))
    check("no near-duplicate names", not near, f"check: {', '.join(near[:5])}", warn=True)

    # 5. sector vocabulary
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from make_workbook import INDUSTRIES
        allowed = set(INDUSTRIES.get(industry, {}).get("sectors") or [])
    except Exception:
        allowed = set()
    if allowed:
        bad = sorted({str(r.get("source_sector_taxonomy")) for r in rows
                      if r.get("source_sector_taxonomy") and str(r["source_sector_taxonomy"]) not in allowed})
        check("sector vocabulary", not bad, f"off-taxonomy: {', '.join(bad[:5])}")

    # 6. distribution sanity
    epis = [float(r["entry_potential_index"]) for r in rows if r.get("entry_potential_index") is not None]
    if epis:
        tiers = Counter(band(e) for e in epis)
        quads = Counter(r.get("quadrant") for r in rows if r.get("quadrant"))
        spread = max(epis) - min(epis)
        print(f"       EPI min {min(epis):.0f} / mean {sum(epis)/len(epis):.0f} / max {max(epis):.0f}")
        print(f"       tiers {dict(tiers)}  quadrants {dict(quads)}")
        ok_spread = spread >= 25
        check("EPI spread is meaningful", ok_spread,
              f"{spread:.0f} pts" + ("" if ok_spread else " — scores may be compressed"), warn=True)
        ok_tiers = tiers.get("High", 0) <= len(epis) * 0.6
        check("not everything is High tier", ok_tiers,
              f"{tiers.get('High', 0)}/{len(epis)} High" + ("" if ok_tiers else " — check for score inflation"), warn=True)

    # 7. anchor drift
    anchors_file = ROOT / "industries" / industry / "anchors.md"
    if anchors_file.exists():
        recorded = dict(re.findall(r"^\|\s*\*{0,2}([^|*]+?)\*{0,2}\s*\|.*?\|\s*(\d+)\s*\|", anchors_file.read_text(encoding="utf-8"), re.M))
        moved = []
        for name, epi in recorded.items():
            row = next((r for r in rows if str(r["company"]).strip().lower() == name.strip().lower()), None)
            if row and row.get("entry_potential_index") is not None:
                if abs(float(row["entry_potential_index"]) - float(epi)) > ANCHOR_TOLERANCE:
                    moved.append(f"{name}: {epi}→{row['entry_potential_index']}")
        check("calibration anchors stable", not moved, "; ".join(moved))
    else:
        check("anchors.md present", False, "record calibration anchors after batch 1", warn=True)

    fails = [r for r in results if r[0] == "FAIL"]
    warns = [r for r in results if r[0] == "WARN"]
    print(f"\n=== {len(results) - len(fails) - len(warns)} passed, {len(warns)} warnings, {len(fails)} failures ===")
    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
