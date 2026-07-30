"""Write a researched batch into an industry workbook, computing all derived scores.

    python -X utf8 scripts/populate.py igaming batches/igaming-batch-01.json
    python -X utf8 scripts/populate.py igaming batches/igaming-batch-01.json --dry-run

Derived scores (W%/D%/AI%/R/EPI/tier/quadrant/financial health) are computed HERE
in Python and written as PLAIN VALUES — never Excel formulas (decision D5): openpyxl
cannot read uncomputed formula caches, which would silently empty the dataset.

Idempotent: a company already present is UPDATED in place, never duplicated.
"""
import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
HEADER_ROW = 3
DATA_START_ROW = 4

# Model Config weights (must match the workbook's Model Config sheet + template DATA-CONTRACT)
WW = [.25, .25, .20, .20, .10]
DW = [.25, .25, .15, .15, .10, .10]
AIW = [.20, .25, .15, .10, .10, .15, .05]
FW = [.20, .20, .20, .20, .20]


def snake(s):
    """Identical to the extractor's header→key rule (template/scripts/extract_excel.py)."""
    s = str(s).strip().replace("%", "pct").replace("&", "and").replace("#", "num")
    s = re.sub(r"[^\w]+", "_", s, flags=re.ASCII)
    return re.sub(r"_+", "_", s).strip("_").lower()


def pct(values, weights):
    """Weighted 0-5 signals → 0-100, renormalized over non-null entries."""
    num = den = 0.0
    for v, w in zip(values or [], weights):
        if v is None:
            continue
        num += float(v) * w
        den += w
    return None if den == 0 else round(num / den / 5 * 100)


def band(epi):
    return None if epi is None else ("High" if epi >= 60 else "Medium" if epi >= 40 else "Low")


def quadrant(w, r):
    if w is None or r is None:
        return None
    return ("Imminent threat" if r >= 60 else "Aspirant") if w >= 60 else ("Sleeping giant" if r >= 60 else "Dormant")


def yn(v):
    if v is None:
        return None
    v = str(v).strip().lower()
    return "y" if v in ("y", "yes", "true", "1") else ("n" if v in ("n", "no", "false", "0") else None)


def derive(row):
    """All computed columns for one researched company."""
    w = pct(row.get("w"), WW)
    d = pct(row.get("d"), DW)
    ai = pct(row.get("ai"), AIW)
    fin = pct(row.get("f"), FW)
    r = None if (d is None or ai is None) else round(.5 * d + .5 * ai)
    epi = None if (w is None or r is None) else round(.4 * w + .6 * r)
    return {"willingness_pct": w, "distribution_readiness_pct": d, "ai_readiness_pct": ai,
            "readiness_pct_combined": r, "entry_potential_index": epi,
            "entry_tier": band(epi), "threat_tier": band(epi), "quadrant": quadrant(w, r),
            "financial_health_pct": fin}


def signal_notes(row):
    """Fold the per-score justifications into one auditable evidence string."""
    out = []
    for prefix, key in (("W", "w"), ("D", "d"), ("AI", "ai"), ("F", "f")):
        scores, notes = row.get(key) or [], row.get(f"{key}_notes") or []
        for i, note in enumerate(notes):
            if note:
                score = scores[i] if i < len(scores) else "?"
                out.append(f"{prefix}{i+1}={score}: {note}")
    return " | ".join(out)


def build_field_map(row, headers):
    """JSON row → {snake_column: value}. Unknown columns are ignored by the writer."""
    vals = {}
    direct = ["hq", "founded", "ftes", "total_funding_mn", "market_cap_valuation_mn",
              "business_model_notes", "customer_base_segment", "revenue_traction", "survival_tier",
              "licensed_jurisdictions", "regulated_revenue_pct", "responsible_gambling_stance",
              "foundation_model_approach", "agentic_maturity_layer", "proprietary_data_advantage",
              "gen_ai_platform_partnerships", "existing_partnerships", "horizon",
              "impact_on_incumbent_line", "evidence_notes", "confidence"]
    for k in direct:
        if row.get(k) is not None:
            vals[k] = row[k]
    vals["company"] = row.get("name")
    vals["source_sector_taxonomy"] = row.get("sector")
    vals["residual_gap_what_they_d_need"] = row.get("residual_gap")
    for k in ("merchant_of_record", "direct_channel", "api_commerce_ready", "mcp", "a2a", "ucp",
              "acp", "crypto_accepted", "is_ai_the_core_focus", "customer_facing_ai_assistant"):
        if row.get(k) is not None:
            vals[k] = yn(row[k])
    if row.get("evidence_links"):
        vals["evidence_links"] = " ; ".join(row["evidence_links"])
    if row.get("notes"):
        vals["team_notes"] = row["notes"]
    notes = signal_notes(row)
    if notes:
        vals["evidence_notes"] = ((vals.get("evidence_notes", "") + " || ") if vals.get("evidence_notes") else "") + notes
    vals["researched_on"] = date.today().isoformat()
    vals["researched_by"] = row.get("_researched_by", "sonnet/company-researcher")
    vals["last_reviewed"] = date.today().isoformat()

    # signal arrays → prefixed columns (w1_*, d1_*, ai1_*, f1_*)
    for key, prefix in (("w", "w"), ("d", "d"), ("ai", "ai"), ("f", "f")):
        for i, v in enumerate(row.get(key) or [], 1):
            col = next((h for h in headers if h.startswith(f"{prefix}{i}_")), None)
            if col:
                vals[col] = v

    # coverage: verticals → "<vertical>_distributing", value chain → vc1..vc7
    covered = {str(v).strip().lower() for v in (row.get("verticals") or [])}
    n_vert = 0
    for h in headers:
        if h.endswith("_distributing"):
            label = h[:-len("_distributing")].replace("_", " ")
            hit = any(label in c or c in label for c in covered)
            vals[h] = "y" if hit else "n"
            n_vert += 1 if hit else 0
    vals["num_verticals_distributed"] = n_vert
    stages = {str(s).strip().lower() for s in (row.get("value_chain") or [])}
    for h in headers:
        m = re.match(r"^vc(\d)_(.+)$", h)
        if m:
            label = m.group(2).replace("_", " ")
            vals[h] = "y" if any(label in s or s in label for s in stages) else "n"

    vals.update(derive(row))
    return {k: v for k, v in vals.items() if v is not None}


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    industry, batch_path = sys.argv[1], Path(sys.argv[2])
    dry = "--dry-run" in sys.argv
    if not batch_path.is_absolute():
        batch_path = ROOT / batch_path
    payload = json.loads(batch_path.read_text(encoding="utf-8"))
    rows = payload.get("rows", payload if isinstance(payload, list) else [])
    if not rows:
        sys.exit("no rows in batch file")

    wb_path = ROOT / "industries" / industry / f"{industry}_landscape.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    ws = wb["Landscape"]
    headers = [snake(c.value) if c.value else None for c in ws[HEADER_ROW]]
    col_of = {h: i + 1 for i, h in enumerate(headers) if h}
    existing = {}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name:
            existing[str(name).strip().lower()] = r

    added = updated = skipped = 0
    unknown_cols = set()
    for row in rows:
        name = (row.get("name") or "").strip()
        if not name:
            skipped += 1
            continue
        vals = build_field_map(row, [h for h in headers if h])
        target = existing.get(name.lower())
        if target is None:
            target = max([DATA_START_ROW - 1] + list(existing.values())) + 1
            existing[name.lower()] = target
            added += 1
        else:
            updated += 1
        for key, value in vals.items():
            if key in col_of:
                if not dry:
                    ws.cell(row=target, column=col_of[key], value=value)
            else:
                unknown_cols.add(key)

    if unknown_cols:
        print(f"note: no column for {sorted(unknown_cols)} (ignored)")
    if dry:
        print(f"DRY RUN — would add {added}, update {updated}, skip {skipped}")
        return
    wb.save(wb_path)

    ledger = ROOT / "industries" / industry / "batches" / "ledger.md"
    ledger.parent.mkdir(parents=True, exist_ok=True)
    line = f"- {date.today().isoformat()} · `{batch_path.name}` · +{added} new, {updated} updated · total now {len(existing)}\n"
    with ledger.open("a", encoding="utf-8") as fh:
        if ledger.stat().st_size == 0:
            fh.write(f"# {industry} population ledger\n\n")
        fh.write(line)
    print(f"{wb_path.relative_to(ROOT)}: +{added} new, {updated} updated, {skipped} skipped → {len(existing)} companies total")


if __name__ == "__main__":
    main()
