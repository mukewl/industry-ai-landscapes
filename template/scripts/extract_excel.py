# ============================================================
# TEMPLATE REFERENCE COPY (from the Amadeus travel-distribution project).
# This is WORKING code for the travel industry, shipped as the reference
# implementation. To adapt it to a new industry, follow the numbered steps
# in ../ADAPTATION-CHECKLIST.md (or ADAPTATION-CHECKLIST.md at template root)
# -- every industry-specific marker in this file is listed there by name.
# ============================================================
"""Extract the Amadeus landscape workbook into data/*.json.

Re-runnable: run whenever the Excel changes (CLAUDE.md staleness check).
    python -X utf8 scripts/extract_excel.py
"""
import hashlib
import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = ROOT / "travel_distribution_landscape_v2_5verticals (2).xlsx"
DATA_DIR = ROOT / "data"

# Landscape sheet layout (verified 2026-06-12)
HEADER_ROW = 3
DATA_START_ROW = 4


def snake(s):
    s = str(s).strip()
    s = s.replace("%", "pct").replace("&", "and").replace("#", "num")
    s = re.sub(r"[^\w]+", "_", s, flags=re.ASCII)
    return re.sub(r"_+", "_", s).strip("_").lower()


def clean(v):
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, date):
        return v.isoformat()
    return v


def extract_landscape(wb):
    ws = wb["Landscape"]
    rows = ws.iter_rows(values_only=True)
    grid = list(rows)
    headers = [snake(h) if h is not None else f"col_{i+1}" for i, h in enumerate(grid[HEADER_ROW - 1])]
    # disambiguate duplicate headers (e.g. total funding appears in 1A and 5A)
    seen = {}
    for i, h in enumerate(headers):
        if h in seen:
            headers[i] = f"{h}_{seen[h] + 1}"
        seen[h] = seen.get(h, 0) + 1

    companies = []
    skipped = []
    for r in grid[DATA_START_ROW - 1:]:
        if not r or clean(r[0]) is None:
            continue
        rec = {headers[j]: clean(r[j]) for j in range(min(len(headers), len(r)))}
        # rows flagged DUPLICATE in reviewer notes are excluded from analysis
        if str(rec.get("reviewer_notes") or "").upper().startswith("DUPLICATE"):
            skipped.append(rec["company"])
            continue
        companies.append(rec)
    if skipped:
        print(f"skipped duplicates: {skipped}")
    return headers, companies


def extract_simple_sheet(wb, name):
    ws = wb[name]
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [clean(v) for v in row]
        if any(c is not None for c in cells):
            out.append(cells)
    return out


def main():
    if not EXCEL_FILE.exists():
        sys.exit(f"Excel file not found: {EXCEL_FILE}")
    DATA_DIR.mkdir(exist_ok=True)

    digest = hashlib.sha256(EXCEL_FILE.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    headers, companies = extract_landscape(wb)
    (DATA_DIR / "companies.json").write_text(
        json.dumps(companies, indent=1, ensure_ascii=False), encoding="utf-8"
    )

    support = {}
    for name in wb.sheetnames:
        if name != "Landscape":
            support[name] = extract_simple_sheet(wb, name)
    (DATA_DIR / "support_sheets.json").write_text(
        json.dumps(support, indent=1, ensure_ascii=False), encoding="utf-8"
    )

    meta = {
        "source_file": EXCEL_FILE.name,
        "sha256": digest,
        "extracted_on": date.today().isoformat(),
        "company_count": len(companies),
        "column_count": len(headers),
        "columns": headers,
    }
    (DATA_DIR / "extract_meta.json").write_text(
        json.dumps(meta, indent=1, ensure_ascii=False), encoding="utf-8"
    )

    print(f"companies.json: {len(companies)} records, {len(headers)} columns")
    print(f"support_sheets.json: {list(support)}")
    print(f"sha256: {digest[:16]}...")


if __name__ == "__main__":
    main()
